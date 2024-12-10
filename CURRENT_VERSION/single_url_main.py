import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from collections import defaultdict
import pandas as pd

from playwright.async_api import async_playwright
import asyncio
import json
import logging
from checks.WCAG_1_3_1.test_blockquote_markup import test_blockquote_markup, write_blockquote_info
from checks.WCAG_1_3_1.test_form_markup import test_form_markup, write_form_info
from checks.WCAG_1_3_1.test_heading_markup import check_heading_markup, write_heading_info
from checks.WCAG_1_3_1.test_landmark_markup import test_landmark_markup, write_landmark_info
from checks.WCAG_1_3_1.test_list_markup import test_list_markup, write_list_info
from checks.WCAG_1_3_1.test_structural_markup import test_structural_markup, write_structural_info
from checks.WCAG_1_3_1.test_table_markup import test_table_markup, write_table_info

def create_results_workbook():
    """Initialize an Excel workbook with the desired column format."""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append([
        "URL",
        "Test Name",
        "Pass/Fail/N/A",
        "Confidence (%)",
        "Issue Details",
    ])
    return workbook

def save_json(file_path, data):

    try:
        with open(file_path, "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, indent=4)
        logging.info(f"JSON file saved: {file_path}")
        return True
    except Exception as e:
        logging.error(f"Error saving JSON to {file_path}: {e}")
        return False


def add_section_header(sheet, test_name):
    """Adds a greyed-out section header for a test type."""
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    header_row = [None] * 5  # Adjust based on the number of columns
    header_row[1] = test_name
    row_idx = sheet.max_row + 1
    sheet.append(header_row)
    for cell in sheet[row_idx]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font


async def fetch_html_content(url):
    """Fetch HTML content from the given URL."""
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            await page.goto(url)
            content = await page.content()
            await browser.close()
            return content
    except Exception as e:
        raise RuntimeError(f"Error fetching HTML content for {url}: {e}")


def save_results(base_dir, test_name, results):
    """
    Saves results in CSV and JSON formats.

    Args:
        base_dir (str): Directory to save results.
        test_name (str): Name of the test.
        results (list): List of dictionaries representing test results.
    """
    try:
        os.makedirs(base_dir, exist_ok=True)
        csv_path = os.path.join(base_dir, f"{test_name}_details.csv")
        json_path = os.path.join(base_dir, f"{test_name}_details.json")

        # Save to CSV
        pd.DataFrame(results).to_csv(csv_path, index=False)

        # Save to JSON
        with open(json_path, "w", encoding="utf-8") as json_file:
            json.dump(results, json_file, indent=4)

        print(f"Results saved for {test_name}: {csv_path}, {json_path}")
    except Exception as e:
        print(f"Failed to save results for {test_name}: {e}")


def group_issues(details):
    """
    Groups issues by their `Issue` field and returns a list of dictionaries with counts.
    
    Args:
        details (list): List of issue dictionaries.

    Returns:
        list: Grouped issues with counts.
    """
    if not details:
        return []

    grouped = defaultdict(list)
    for detail in details:
        issue_key = detail.get("Issue", "No issue description provided.")
        grouped[issue_key].append(detail)

    # Convert grouped issues to list format with counts
    return [
        {**occurrences[0], "Count": len(occurrences)}
        for occurrences in grouped.values()
    ]


def process_url(url, workbook, results_dir):
    """
    Processes the given URL, performs tests, and saves results in separate folders for each test.

    Args:
        url (str): URL to process.
        workbook (Workbook): Excel workbook for results.
        results_dir (str): Directory to save results.

    Returns:
        str: Path to the summary Excel file.
    """
    try:
        # Fetch HTML content
        html_content = asyncio.run(fetch_html_content(url))
        if not html_content:
            print(f"Failed to fetch HTML content for {url}.")
            return

        summary_sheet = workbook["Summary"]

        # Define tests and their associated functions
        tests = {
            "Heading Markup": (check_heading_markup, write_heading_info),
            "List Markup": (test_list_markup, write_list_info),
            "Table Markup": (test_table_markup, write_table_info),
            "Blockquote Markup": (test_blockquote_markup, write_blockquote_info),
            "Landmark Markup": (test_landmark_markup, write_landmark_info),
            "Structural Markup": (test_structural_markup, write_structural_info),
            "Form Markup": (test_form_markup, write_form_info),
        }

        # Create folders for each test before running them
        test_folders = {}
        for test_name in tests.keys():
            test_folder = os.path.join(results_dir, test_name.replace(" ", "_"))
            os.makedirs(test_folder, exist_ok=True)
            test_folders[test_name] = test_folder

        # Run each test and save results
        for test_name, (test_function, write_function) in tests.items():
            try:
                # Add section header in Excel
                add_section_header(summary_sheet, test_name)

                # Run the test
                result = test_function(html_content)
                logging.debug(f"Test Result for {test_name}: {result}")

                # Extract test results
                status = result.get("status", "N/A")
                confidence = result.get("confidence", 100.0)
                details = result.get("details", [])

                # Group issues for better organization
                grouped_results = group_issues(details)

                # Add grouped results to the Excel summary
                if grouped_results:
                    for item in grouped_results:
                        summary_sheet.append([
                            url,
                            test_name,
                            status,
                            f"{confidence:.2f}%",
                            item["Issue"] + f" (Occurred {item['Count']} times)"
                        ])
                else:
                    summary_sheet.append([url, test_name, status, f"{confidence:.2f}%", "No issues found."])

                # Save results to test-specific folder
                test_folder = test_folders[test_name]
                csv_file_path = os.path.join(test_folder, f"{test_name.replace(' ', '_')}_details.csv")
                json_file_path = os.path.join(test_folder, f"{test_name.replace(' ', '_')}_details.json")

                try:
                    # Save CSV and JSON
                    write_function(csv_file_path, grouped_results)
                    save_json(json_file_path, grouped_results)
                    logging.info(f"Saved results for {test_name} in {test_folder}")
                except Exception as file_write_error:
                    logging.error(f"Error saving results for {test_name}: {file_write_error}")

            except Exception as e:
                # Log and record test-specific errors
                error_message = f"Error processing {test_name} for {url}: {e}"
                logging.error(error_message)
                summary_sheet.append([url, test_name, "Error", "0.00%", error_message])

        # Save the Excel summary file
        excel_path = os.path.join(results_dir, "wcag1.3.1_summary.xlsx")
        workbook.save(excel_path)
        return excel_path

    except Exception as e:
        logging.error(f"Unexpected error processing {url}: {e}")



def open_results_file(file_path):
    """Open the results file in the default application."""
    try:
        if os.name == "nt":
            os.startfile(file_path)
        elif os.name == "posix":
            subprocess.run(["open" if "darwin" in os.uname().sysname.lower() else "xdg-open", file_path])
        else:
            print(f"Cannot determine how to open files on this platform.")
    except Exception as e:
        print(f"Failed to open the file {file_path}: {e}")

def main():
    url = None
    while not url:
        url = input("Enter the URL to test: ").strip()
        if not url:
            print("No URL provided. Please enter a valid URL.")

    base_results_dir = "output"
    os.makedirs(base_results_dir, exist_ok=True)

    url_safe_name = url.replace("http://", "").replace("https://", "").replace("/", "_")
    url_results_dir = os.path.join(base_results_dir, url_safe_name)
    os.makedirs(url_results_dir, exist_ok=True)

    workbook = create_results_workbook()

    print(f"\nTesting URL: {url}")
    results_file = process_url(url, workbook, url_results_dir)

    if results_file:
        open_results_file(results_file)

    print(f"\nTest completed. Results saved in {results_file}.")


if __name__ == "__main__":
    main()